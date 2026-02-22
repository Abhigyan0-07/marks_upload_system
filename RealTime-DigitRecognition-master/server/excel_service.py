from openpyxl import Workbook, load_workbook
import io
import os
from datetime import datetime

class ExcelService:
    @staticmethod
    def create_or_append_marks(file_path, marks_list):
        """Appends a new row of marks to an Excel file and sums them."""
        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            # Header
            headers = ["Timestamp"] + [f"Q{i+1}" for i in range(len(marks_list))] + ["Total"]
            ws.append(headers)
        else:
            wb = load_workbook(file_path)
            ws = wb.active

        # Prepare new row
        total = sum(marks_list)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row = [timestamp] + marks_list + [total]
        ws.append(row)
        
        wb.save(file_path)
        return total

    @staticmethod
    def get_grand_total(file_path):
        """Calculates the sum of all 'Total' columns in the file."""
        if not os.path.exists(file_path):
            return 0
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        grand_total = 0
        # Assuming header is row 1, data starts row 2. Total is last column.
        total_col = ws.max_column
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=total_col).value
            if isinstance(val, (int, float)):
                grand_total += val
        
        return grand_total
